from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
import openpyxl
import time
import datetime


class DataCatch:
    def __init__(self, navegador):
        self.__navegador = navegador


    def PesquisaCep(self, cep):
        self.__navegador.find_element(by=By.ID, value="endereco").send_keys(cep)
        time.sleep(2)
        self.__navegador.find_element(by=By.ID, value="btn_pesquisar").click()
        time.sleep(2)


    def CapturaDados(self):
        logradouro = self.__navegador.find_element('xpath', '//*[@id="resultado-DNEC"]/tbody/tr/td[1]').accessible_name
        bairro = self.__navegador.find_element('xpath', '//*[@id="resultado-DNEC"]/tbody/tr/td[2]').accessible_name
        localidade = self.__navegador.find_element('xpath', '//*[@id="resultado-DNEC"]/tbody/tr/td[3]').accessible_name
        cep = self.__navegador.find_element('xpath', '//*[@id="resultado-DNEC"]/tbody/tr[1]/td[4]').accessible_name

        self.__navegador.find_element(by=By.ID, value="btn_nbusca").click()
        time.sleep(3)
        return [cep,logradouro,bairro,localidade]
    

    def ReadPlanilha():
        planilha = load_workbook("C:/Users/leona/Documents/projetos/extracao_cep/input/" + "cep.xlsx")
        sheet = planilha.active
        return sheet["A:A"]
    

    def EscreveDados(dados):
        data = datetime.datetime.now().strftime("%Y-%m-%d")
        output = "C:/Users/leona/Documents/projetos/extracao_cep/output/" + "output_cep_" + str(data) + ".xlsx"

        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        for linha in dados:
            worksheet.append(linha)
        
        workbook.save(output)